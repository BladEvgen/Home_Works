import sys
import datetime
import os
import sqlite3
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from PyQt6.QtWidgets import (
    QMessageBox,
    QMainWindow,
    QApplication,
    QLineEdit,
    QTextEdit,
)
from PyQt6 import uic
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.header import Header
import smtplib
from dotenv import load_dotenv

dotenv_path = os.path.join(os.path.dirname(__file__), ".env")
if os.path.exists(dotenv_path):
    load_dotenv(dotenv_path)


class DatabaseSQLite:
    def __init__(self, database_name: str):
        self.database_name = database_name
        self.connection = None

    def connect(self):
        self.connection = sqlite3.connect(self.database_name)

    def create_tables(self):
        query_items = """
        CREATE TABLE IF NOT EXISTS items 
        (
        id INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL,
        number INTEGER NOT NULL,
        price REAL DEFAULT '0.0'
        )
        """
        query_candidates = """
        CREATE TABLE IF NOT EXISTS candidates 
        (
        id INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL,
        name VARCHAR NOT NULL,
        role VARCHAR NOT NULL
        )
        """
        with self.connection:
            self.connection.execute(query_items)
            self.connection.execute(query_candidates)

    def insert_item(self, number: int, price: float):
        query = """
        INSERT INTO items (number, price)
        VALUES (?, ?)
        """
        with self.connection:
            self.connection.execute(query, (number, price))

    def get_all_items(self):
        query = """
        SELECT * FROM items;
        """
        with self.connection:
            cursor = self.connection.execute(query)
            rows = cursor.fetchall()
        return rows

    def insert_candidate(self, name: str, role: str):
        query = """
        INSERT INTO candidates (name, role)
        VALUES (?, ?)
        """
        with self.connection:
            self.connection.execute(query, (name, role))

    def get_all_candidates(self):
        query = """
        SELECT * FROM candidates;
        """
        with self.connection:
            cursor = self.connection.execute(query)
            rows = cursor.fetchall()
        return rows


class ExcelExporter:
    def export_to_excel(self, data, filename):
        workbook = Workbook()
        worksheet = workbook.active

        headers = ["ИИН", "Номер заявки", "Сумма заявки"]
        for col, header in enumerate(headers, start=1):
            worksheet.cell(row=1, column=col, value=header)

        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                worksheet.cell(row=row_idx, column=col_idx, value=value)

        workbook.save(filename)


class EmailSender:
    def send_email(self, recipients, subject, message, candidates_data):
        recipients_email = recipients.split(",")
        try:
            login = os.getenv("login")
            password = os.getenv("password")
            msg = MIMEMultipart()
            msg["Subject"] = Header(subject, "utf-8")
            msg["From"] = login
            msg["To"] = ", ".join(recipients_email)

            html_table = "<table border='1'><tr><th>ID</th><th>Имя кандидата</th><th>Роль кандидата</th></tr>"
            for row in candidates_data:
                html_table += (
                    f"<tr><td>{row[0]}</td><td>{row[1]}</td><td>{row[2]}</td></tr>"
                )
            html_table += "</table>"

            msg.attach(MIMEText(html_table, "html", "utf-8"))
            msg.attach(MIMEText(message, "plain", "utf-8"))

            s = smtplib.SMTP("smtp.gmail.com", 587, timeout=10)
            try:
                s.starttls()
                s.login(login, password)
                s.sendmail(msg["From"], recipients_email, msg.as_string())
            except Exception as ex:
                print(ex)
            finally:
                s.quit()

            return True
        except Exception as error:
            print(error)
            return False


class Ui(QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = uic.loadUi("main.ui")
        self.setCentralWidget(self.ui)

        self.ui.pushButton_save.clicked.connect(self.save_to_database)
        self.ui.pushButton_export.clicked.connect(self.export_from_database)
        self.ui.pushButton_save_candidate.clicked.connect(
            self.save_candidate_to_database
        )
        self.ui.pushButton_export_candidates.clicked.connect(
            self.export_candidates_to_excel
        )
        self.ui.pushButton_send_email.clicked.connect(self.send_email)

        self.setMinimumSize(400, 500)
        self.setMaximumSize(400, 500)
        self.show()

        self.lineEdit_email = self.ui.findChild(QLineEdit, "lineEdit_email")
        self.textEdit_email_message = self.ui.findChild(
            QTextEdit, "textEdit_email_message"
        )
        self.lineEdit_number = self.ui.findChild(QLineEdit, "lineEdit_number")
        self.lineEdit_candidate_name = self.ui.findChild(
            QLineEdit, "lineEdit_candidate_name"
        )

        self.database = DatabaseSQLite("database_items.db")
        self.database.connect()
        self.database.create_tables()

    def save_to_database(self):
        try:
            number = self.lineEdit_number.text().strip()
            is_number = self.validate_number(value=number)
            if not is_number:
                QMessageBox.critical(
                    self, "Ошибка", "Вы ввели неправильный номер заявки"
                )
                return

            price = self.ui.doubleSpinBox_price.value()
            self.database.insert_item(number, price)
            QMessageBox.information(
                self, "Успешно", "Заявка успешно добавлена в базу данных"
            )
        except Exception as error:
            print(error)

    def export_from_database(self):
        try:
            rows = self.database.get_all_items()
            excel_exporter = ExcelExporter()
            filename = f"экспорт_{datetime.datetime.now().strftime('%Y-%m-%d')}.xlsx"
            excel_exporter.export_to_excel(rows, filename)
            QMessageBox.information(self, "Успешно", "Данные успешно экспортированы")
        except Exception as error:
            print(error)

    def validate_number(self, value: str) -> bool:
        try:
            value = int(value)
            if value < 1:
                return False
        except Exception as error:
            print("Error: ", error)
            return False
        return True

    def save_candidate_to_database(self):
        try:
            name = self.lineEdit_candidate_name.text().strip()
            role = self.ui.lineEdit_candidate_role.text().strip()

            if not name or not role:
                QMessageBox.critical(self, "Ошибка", "Заполните все поля для кандидата")
                return

            self.database.insert_candidate(name, role)
            QMessageBox.information(
                self, "Успешно", "Кандидат успешно добавлен в базу данных"
            )
        except Exception as error:
            print(error)

    def export_candidates_to_excel(self):
        try:
            rows = self.database.get_all_candidates()
            excel_exporter = ExcelExporter()
            filename = f"экспорт_кандидатов_{datetime.datetime.now().strftime('%Y-%m-%d')}.xlsx"
            excel_exporter.export_to_excel(rows, filename)
            QMessageBox.information(
                self, "Успешно", "Данные кандидатов успешно экспортированы"
            )
        except Exception as error:
            print(error)

    def send_email(self):
        recipients_email = self.ui.lineEdit_email.text()
        candidates_data = self.database.get_all_candidates()
        email_message = self.ui.textEdit_email_message.toPlainText()

        email_sender = EmailSender()
        success = email_sender.send_email(
            recipients_email, "Важно к прочтении", email_message, candidates_data
        )

        if success:
            QMessageBox.information(self, "Успешно", "Данные успешно отправлены")
        else:
            QMessageBox.critical(self, "Ошибка", "Не удалось отправить письмо")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    ui = Ui()
    sys.exit(app.exec())
