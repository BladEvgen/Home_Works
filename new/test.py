import os
from openpyxl import load_workbook

def merge_data(file1, file2, file3, output_file):
    wb1 = load_workbook(file1)
    wb2 = load_workbook(file2)
    wb3 = load_workbook(file3)

    sheet1 = wb1.active
    sheet2 = wb2.active
    sheet3 = wb3.active

    merged_wb = load_workbook(output_file)
    merged_sheet = merged_wb.active

    # Extract data from the first file and insert into column A of the merged file
    for row in sheet1.iter_rows(min_row=1, values_only=True):
        merged_sheet.append(row)

    # Extract data from the second file and insert into column B of the merged file
    for row in sheet2.iter_rows(min_row=1, values_only=True):
        if row[0] is not None and row[1] is not None:
            merged_sheet.cell(row=row[0], column=2, value=row[1])

    # Extract data from the third file and insert into column C of the merged file
    for row in sheet3.iter_rows(min_row=1, values_only=True):
        if row[0] is not None and row[1] is not None:
            merged_sheet.cell(row=row[0], column=3, value=row[1])

    merged_wb.save(output_file)
    print("Data merged successfully!")

def main():
    file1 = os.path.join(os.path.dirname(__file__), "list1.xlsx")
    file2 = os.path.join(os.path.dirname(__file__), "list2.xlsx")
    file3 = os.path.join(os.path.dirname(__file__), "list3.xlsx")
    merge_file = os.path.join(os.path.dirname(__file__), "merge_file.xlsx")
    merge_data(file1, file2, file3, merge_file)

main()
