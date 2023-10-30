import sys
import sqlite3
import datetime
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from PyQt6.QtWidgets import (
    QMessageBox,
    QGridLayout,
    QTableWidget,
    QTableWidgetItem,
    QMainWindow,
    QApplication,
    QPushButton,
    QLineEdit,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)
from PyQt6 import uic
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.header import Header
import smtplib
import os
from dotenv import load_dotenv

dotenv_path = os.path.join(os.path.dirname(__file__), ".env")
if os.path.exists(dotenv_path):
    load_dotenv(dotenv_path)


class Ui(QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = uic.loadUi("main.ui")
        self.setCentralWidget(self.ui)

        self.ui.pushButton_save.clicked.connect(self.save_to_database)
        self.ui.pushButton_export.clicked.connect(self.export_from_database)
        self.ui.pushButton_save_candidate.clicked.connect(
            self.save_candidate_to_database
        )
        self.ui.pushButton_export_candidates.clicked.connect(
            self.export_candidates_to_excel
        )
        self.ui.pushButton_send_email.clicked.connect(self.send_email)

        self.setMinimumSize(400, 500)
        self.setMaximumSize(400,500)
        self.show()

        self.lineEdit_email = self.ui.findChild(QLineEdit, "lineEdit_email")
        self.textEdit_email_message = self.ui.findChild(
            QTextEdit, "textEdit_email_message"
        )
        self.lineEdit_number = self.ui.findChild(QLineEdit, "lineEdit_number")
        self.lineEdit_candidate_name = self.ui.findChild(
            QLineEdit, "lineEdit_candidate_name"
        )

    def save_to_database(self):
        try:
            number = self.lineEdit_number.text().strip()
            is_number = self.validate_number(value=number)
            if not is_number:
                QMessageBox.critical(
                    self, "Ошибка", "Вы ввели неправильный номер заявки"
                )
                return

            price = self.ui.doubleSpinBox_price.value()
            connection = sqlite3.connect("database_items.db")
            cursor = connection.cursor()
            query = """
            INSERT INTO items (number, price)
            VALUES (?, ?)
            """
            cursor.execute(query, (number, price))
            connection.commit()
            QMessageBox.information(
                self, "Успешно", "Заявка успешно добавлена в базу данных"
            )
        except Exception as error:
            print(error)

    def export_from_database(self):
        try:
            rows = self.get_all_items()
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            for column_i, column in enumerate(
                ["ИИН", "Номер заявки", "Сумма заявки"], 1
            ):
                worksheet.cell(row=1, column=column_i, value=column)

            for row_i, row in enumerate(rows, 2):
                for column_i, column in enumerate(row, 1):
                    worksheet.cell(row=row_i, column=column_i, value=column)

            workbook.save(
                f"экспорт_{datetime.datetime.now().strftime('%Y-%m-%d')}.xlsx"
            )
            QMessageBox.information(self, "Успешно", "Данные успешно экспортированы")
        except Exception as error:
            print(error)

    def validate_number(self, value: str) -> bool:
        try:
            value = int(value)
            if value < 1:
                return False
        except Exception as error:
            print("Error: ", error)
            return False
        return True

    def get_all_items(self):
        connection = sqlite3.connect("database_items.db")
        cursor = connection.cursor()
        query1 = """
        SELECT * FROM items;
        """
        cursor.execute(query1)
        rows = cursor.fetchall()
        return rows

    def save_candidate_to_database(self):
        try:
            name = self.lineEdit_candidate_name.text().strip()
            role = self.ui.lineEdit_candidate_role.text().strip()

            if not name or not role:
                QMessageBox.critical(self, "Ошибка", "Заполните все поля для кандидата")
                return

            connection = sqlite3.connect("database_items.db")
            cursor = connection.cursor()
            query = """
            INSERT INTO candidates (name, role)
            VALUES (?, ?)
            """
            cursor.execute(query, (name, role))
            connection.commit()
            QMessageBox.information(
                self, "Успешно", "Кандидат успешно добавлен в базу данных"
            )
        except Exception as error:
            print(error)

    def export_candidates_to_excel(self):
        try:
            rows = self.get_all_candidates()
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            for column_i, column in enumerate(
                ["ID", "Имя кандидата", "Роль кандидата"], 1
            ):
                worksheet.cell(row=1, column=column_i, value=column)

            for row_i, row in enumerate(rows, 2):
                for column_i, column in enumerate(row, 1):
                    worksheet.cell(row=row_i, column=column_i, value=column)

            workbook.save(
                f"экспорт_кандидатов_{datetime.datetime.now().strftime('%Y-%m-%d')}.xlsx"
            )
            QMessageBox.information(
                self, "Успешно", "Данные кандидатов успешно экспортированы"
            )
        except Exception as error:
            print(error)

    def get_all_candidates(self):
        connection = sqlite3.connect("database_items.db")
        cursor = connection.cursor()
        query = """
        SELECT * FROM candidates;
        """
        cursor.execute(query)
        rows = cursor.fetchall()
        return rows

    def send_email(self):
        recipients_email = self.ui.lineEdit_email.text().split(",")
        candidates_data = self.get_all_candidates()
        email_message = self.ui.textEdit_email_message.toPlainText()
    
        try:
            login = os.getenv("login")
            password = os.getenv("password")
            msg = MIMEMultipart()
            msg["Subject"] = Header("Важно к прочтении", "utf-8")
            msg["From"] = login
            msg["To"] = ", ".join(recipients_email)
    
            # Создаем HTML-таблицу на основе данных о кандидатах
            html_table = "<table border='1'><tr><th>ID</th><th>Имя кандидата</th><th>Роль кандидата</th></tr>"
            for row in candidates_data:
                html_table += (
                    f"<tr><td>{row[0]}</td><td>{row[1]}</td><td>{row[2]}</td></tr>"
                )
            html_table += "</table>"
    
            # Добавляем HTML-таблицу и текст сообщения к сообщению
            msg.attach(MIMEText(html_table, "html", "utf-8"))
            msg.attach(MIMEText(email_message, "plain", "utf-8"))
    
            # Отправляем письмо
            s = smtplib.SMTP("smtp.gmail.com", 587, timeout=10)
            try:
                s.starttls()
                s.login(login, password)
                s.sendmail(msg["From"], recipients_email, msg.as_string())
            except Exception as ex:
                print(ex)
            finally:
                s.quit()
    
            QMessageBox.information(self, "Успешно", "Данные успешно отправлены")
        except Exception as error:
            print(error)
    

def database_sqlite():
    connection = sqlite3.connect("database_items.db")
    cursor = connection.cursor()
    query_items = """
    CREATE TABLE IF NOT EXISTS items 
    (
    id INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL,
    number INTEGER NOT NULL,
    price REAL DEFAULT '0.0'
    )
    """
    cursor.execute(query_items)

    query_candidates = """
    CREATE TABLE IF NOT EXISTS candidates 
    (
    id INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL,
    name VARCHAR NOT NULL,
    role VARCHAR NOT NULL
    )
    """
    cursor.execute(query_candidates)

    connection.commit()


if __name__ == "__main__":
    # database_sqlite()
    app = QApplication(sys.argv)
    ui = Ui()
    sys.exit(app.exec())
