import sys
import datetime
import os
import sqlite3
import openpyxl
from openpyxl.workbook import Workbook
from PyQt6.QtWidgets import (
    QMessageBox,
    QMainWindow,
    QApplication,
    QLineEdit,
    QPushButton,
    QDoubleSpinBox,
    QTextEdit,
    QCalendarWidget,
    QDateEdit
)
from PyQt6 import uic


class DatabaseSQLite:
    def __init__(self, database_name: str):
        self.database_name = database_name
        self.connection = None

    def connect(self):
        self.connection = sqlite3.connect(self.database_name)

    def create_tables(self):
        query_products = """
        CREATE TABLE IF NOT EXISTS products 
        (
        id INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL,
        name VARCHAR NOT NULL,
        price REAL DEFAULT '0.0'
        )
        """
        query_production_records = """
        CREATE TABLE IF NOT EXISTS production_records 
        (
        id INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL,
        product_id INTEGER NOT NULL,
        quantity INTEGER NOT NULL,
        date DATE NOT NULL
        )
        """
        with self.connection:
            self.connection.execute(query_products)
            self.connection.execute(query_production_records)

    def insert_product(self, name: str, price: float):
        query = """
        INSERT INTO products (name, price)
        VALUES (?, ?)
        """
        with self.connection:
            self.connection.execute(query, (name, price))

    def get_all_products(self):
        query = """
        SELECT * FROM products;
        """
        with self.connection:
            cursor = self.connection.execute(query)
            rows = cursor.fetchall()
        return rows

    def insert_production_record(self, product_id: int, quantity: int, date: str):
        query = """
        INSERT INTO production_records (product_id, quantity, date)
        VALUES (?, ?, ?)
        """
        with self.connection:
            self.connection.execute(query, (product_id, quantity, date))

    def get_all_production_records(self):
        query = """
        SELECT pr.id, p.name, pr.quantity, pr.date
        FROM production_records pr
        JOIN products p ON pr.product_id = p.id;
        """
        with self.connection:
            cursor = self.connection.execute(query)
            rows = cursor.fetchall()
        return rows


class ExcelExporter:
    def export_to_excel(self, data, filename):
        workbook = Workbook()
        worksheet = workbook.active

        headers = ["ID", "Наименование продукта", "Количество", "Дата (д.м.г)"]
        for col, header in enumerate(headers, start=1):
            worksheet.cell(row=1, column=col, value=header)

        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                worksheet.cell(row=row_idx, column=col_idx, value=value)

        workbook.save(filename)


class Ui(QMainWindow):
    def __init__(self):
        super().__init__()
        uic.loadUi("main.ui", self)

        self.pushButton_save_product.clicked.connect(self.save_product_to_database)
        self.pushButton_export_production.clicked.connect(self.export_production_records_to_excel)

        self.show()
        self.lineEdit_product_name = self.findChild(QLineEdit, "lineEdit_product_name")
        self.doubleSpinBox_product_price = self.findChild(QDoubleSpinBox, "doubleSpinBox_product_price")
        self.lineEdit_quantity = self.findChild(QLineEdit, "lineEdit_quantity")
        self.dateEdit_date = self.findChild(QDateEdit, "dateEdit_date")

        self.database = DatabaseSQLite("production_database.db")
        self.database.connect()
        self.database.create_tables()

    def save_product_to_database(self):
        try:
            name = self.lineEdit_product_name.text().strip()
            price = self.doubleSpinBox_product_price.value()
            quantity = int(self.lineEdit_quantity.text())
            date = self.dateEdit_date.date().toString("dd-MM-yyyy")

            self.database.insert_product(name, price)
            product_id = self.get_last_inserted_product_id()
            self.database.insert_production_record(product_id, quantity, date)

            QMessageBox.information(self, "Успешно", "Продукт успешно добавлен и запись о производстве сохранена")
        except Exception as error:
            print(error)

    def get_last_inserted_product_id(self):
        query = "SELECT last_insert_rowid();"
        with self.database.connection:
            cursor = self.database.connection.execute(query)
            row = cursor.fetchone()
            if row:
                return row[0]
            return None

    def export_production_records_to_excel(self):
        try:
            rows = self.database.get_all_production_records()
            excel_exporter = ExcelExporter()
            filename = f"экспорт_производства_{datetime.datetime.now().strftime('%Y-%m-%d')}.xlsx"
            excel_exporter.export_to_excel(rows, filename)
            QMessageBox.information(
                self, "Успешно", "Данные о производстве успешно экспортированы"
            )
        except Exception as error:
            print(error)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    ui = Ui()
    sys.exit(app.exec())
