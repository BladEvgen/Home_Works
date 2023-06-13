import openpyxl
from openpyxl.utils import get_column_letter
import os


def merge_xlsx_files(file1, file2, file3, merge_file):
    wb_merge = openpyxl.Workbook()
    ws_merge = wb_merge.active

    wb1 = openpyxl.load_workbook(file1)
    ws1 = wb1.active

    wb2 = openpyxl.load_workbook(file2)
    ws2 = wb2.active

    wb3 = openpyxl.load_workbook(file3)
    ws3 = wb3.active

    max_rows = max(ws1.max_row, ws2.max_row, ws3.max_row)
    max_columns = max(ws1.max_column, ws2.max_column, ws3.max_column)

    for row in range(1, max_rows + 1):
        merged_row = []
        for col in range(1, max_columns + 1):
            value1 = ws1.cell(row=row, column=col).value if col <= ws1.max_column else None
            value2 = ws2.cell(row=row, column=col).value if col <= ws2.max_column else None
            value3 = ws3.cell(row=row, column=col).value if col <= ws3.max_column else None
            merged_row.append(value1)
            merged_row.append(value2)
            merged_row.append(value3)
        ws_merge.append(merged_row)

    # Remove three additional empty cells
    for row in ws_merge.iter_rows():
        row = row[:-(3 if max_columns > 0 else 0)]

    # Save the merged file
    wb_merge.save(merge_file)


    # Save the merged file
    wb_merge.save(merge_file)

    wb_merged = openpyxl.load_workbook(merge_file)
    ws_merged = wb_merged.active

    for row in ws_merged.iter_rows(values_only=True):
        print(row)


def main():
    file1 = os.path.join(os.path.dirname(__file__), "list1.xlsx")
    file2 = os.path.join(os.path.dirname(__file__), "list2.xlsx")
    file3 = os.path.join(os.path.dirname(__file__), "list3.xlsx")
    merge_file = os.path.join(os.path.dirname(__file__), "merge_file.xlsx")
    merge_xlsx_files(file1, file2, file3, merge_file)


if __name__ == "__main__":
    main()
