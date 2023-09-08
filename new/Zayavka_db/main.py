import sys
import sqlite3
import datetime
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from PyQt6.QtWidgets import QApplication, QWidget, QMessageBox, QGridLayout, QPushButton
from PyQt6 import uic


class Ui(QWidget):
    def __init__(self):
        super().__init__()
        self.ui = uic.loadUi("main.ui", self)
        self.ui.pushButton_save.clicked.connect(self.save_to_database)
        self.ui.pushButton_export.clicked.connect(self.export_from_database)
        self.ui.pushButton_save_candidate.clicked.connect(
            self.save_candidate_to_database
        )
        self.ui.pushButton_export_candidates.clicked.connect(
            self.export_candidates_to_excel
        )
        self.setMinimumSize(400, 400)
        self.show()

    def save_to_database(self):
        try:
            number = self.lineEdit_number.text().strip()
            is_number = self.validate_number(value=number)
            if not is_number:
                QMessageBox.critical(
                    self, "Ошибка", "Вы ввели неправильный номер заявки"
                )
                return

            price = self.doubleSpinBox_price.value()
            connection = sqlite3.connect("database_items.db")
            cursor = connection.cursor()
            query = """
            INSERT INTO items (number, price)
            VALUES (?, ?)
            """
            cursor.execute(query, (number, price))
            connection.commit()
            QMessageBox.information(
                self, "Успешно", "Заявка успешно добавлена в базу данных"
            )
        except Exception as error:
            print(error)

    def export_from_database(self):
        try:
            rows = self.get_all_items()
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            for column_i, column in enumerate(
                ["ИИН", "Номер заявки", "Сумма заявки"], 1
            ):
                worksheet.cell(row=1, column=column_i, value=column)

            for row_i, row in enumerate(rows, 2):
                for column_i, column in enumerate(row, 1):
                    worksheet.cell(row=row_i, column=column_i, value=column)

            workbook.save(
                f"экспорт_{datetime.datetime.now().strftime('%Y-%m-%d')}.xlsx"
            )
            QMessageBox.information(self, "Успешно", "Данные успешно экспортированы")
        except Exception as error:
            print(error)

    def validate_number(self, value: str) -> bool:
        try:
            value = int(value)
            if value < 1:
                return False
        except Exception as error:
            print("Error: ", error)
            return False
        return True

    def validate_price(self):
        pass

    def get_all_items(self):
        connection = sqlite3.connect("database_items.db")
        cursor = connection.cursor()
        query1 = """
        SELECT * FROM items;
        """
        cursor.execute(query1)
        rows = cursor.fetchall()
        return rows

    def save_candidate_to_database(self):
        try:
            name = self.lineEdit_candidate_name.text().strip()
            role = self.lineEdit_candidate_role.text().strip()

            if not name or not role:
                QMessageBox.critical(self, "Ошибка", "Заполните все поля для кандидата")
                return

            connection = sqlite3.connect("database_items.db")
            cursor = connection.cursor()
            query = """
            INSERT INTO candidates (name, role)
            VALUES (?, ?)
            """
            cursor.execute(query, (name, role))
            connection.commit()
            QMessageBox.information(
                self, "Успешно", "Кандидат успешно добавлен в базу данных"
            )
        except Exception as error:
            print(error)

    def export_candidates_to_excel(self):
        try:
            rows = self.get_all_candidates()
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            for column_i, column in enumerate(
                ["ID", "Имя кандидата", "Роль кандидата"], 1
            ):
                worksheet.cell(row=1, column=column_i, value=column)

            for row_i, row in enumerate(rows, 2):
                for column_i, column in enumerate(row, 1):
                    worksheet.cell(row=row_i, column=column_i, value=column)

            workbook.save(
                f"экспорт_кандидатов_{datetime.datetime.now().strftime('%Y-%m-%d')}.xlsx"
            )
            QMessageBox.information(
                self, "Успешно", "Данные кандидатов успешно экспортированы"
            )
        except Exception as error:
            print(error)

    def get_all_candidates(self):
        connection = sqlite3.connect("database_items.db")
        cursor = connection.cursor()
        query = """
        SELECT * FROM candidates;
        """
        cursor.execute(query)
        rows = cursor.fetchall()
        return rows


def database_postgre():
    """
    CREATE TABLE items (
    id SERIAL PRIMARY KEY,
    number BIGINT NOT NULL,
    price DECIMAL(10, 2) NOT NULL default = '0.0'
    );
    """


def database_sqlite():
    connection = sqlite3.connect("database_items.db")
    cursor = connection.cursor()
    query_items = """
    CREATE TABLE IF NOT EXISTS items 
    (
    id INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL,
    number INTEGER NOT NULL,
    price REAL DEFAULT '0.0'
    )
    """
    cursor.execute(query_items)

    query_candidates = """
    CREATE TABLE IF NOT EXISTS candidates 
    (
    id INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL,
    name VARCHAR NOT NULL,
    role VARCHAR NOT NULL
    )
    """
    cursor.execute(query_candidates)

    connection.commit()


class Modal(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.resize(400, 400)
        self.setWindowTitle("Вы ввели неверный номер заявки")
        self.grid = QGridLayout(self)
        self.button = QPushButton()
        self.button.setText("ОК")
        self.grid.addWidget(self.button)


if __name__ == "__main__":
    # database_sqlite()
    app = QApplication(sys.argv)
    ui = Ui()
    sys.exit(app.exec())
