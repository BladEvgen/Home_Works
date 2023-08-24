import os
from openpyxl import load_workbook

def total_python(file: str) -> int | float:
    wb = load_workbook(file)
    sheet1 = wb.active
    total_sum = 0

    for row in sheet1.iter_rows(min_row=2, values_only=True):
        for cell_value in row:
            if isinstance(cell_value, (int, float)):
                total_sum += cell_value

    total_value = f"Total: {total_sum}"
    
    sheet1.cell(row=sheet1.max_row + 2, column=1, value=total_value)
    wb.save(file)
    wb.close()
    
    return total_sum

def main():
    file1 = os.path.join(os.path.dirname(__file__), "list1.xlsx")
    total_sum = total_python(file1)
    
    print(f"Сумма всех цифр {total_sum}")

if __name__ == '__main__':
    main()
