import os
from openpyxl import load_workbook, Workbook


def merge_xlsx_data(files, output_file):
    merged_file = Workbook()
    merged_sheet = merged_file.active

    column = 1

    for file in files:
        wb = load_workbook(file)
        sheet = wb.active

        for row in sheet.iter_rows():
            for cell in row:
                merged_sheet.cell(row=cell.row, column=column).value = cell.value

        column += 1

    merged_file.save(output_file)


def main():
    file1 = os.path.join(os.path.dirname(__file__), "list1.xlsx")
    file2 = os.path.join(os.path.dirname(__file__), "list2.xlsx")
    file3 = os.path.join(os.path.dirname(__file__), "list3.xlsx")
    merge_file = os.path.join(os.path.dirname(__file__), "merge_file_ok.xlsx")
    file_paths = [file1, file2, file3]
    merge_xlsx_data(files=file_paths, output_file=merge_file)


if __name__ == "__main__":
    main()
