import sqlite3
from PyQt6.QtCore import QRect
from PyQt6.QtWidgets import (
    QApplication,
    QMainWindow,
    QPushButton,
    QFileDialog,
    QLabel,
    QGridLayout,
    QWidget
)
import sys
from openpyxl import load_workbook, Workbook


def query(query_str: str, args=(), many=True) -> list | None:
    with sqlite3.connect("db/database_1.db") as connection:
        cursor = connection.cursor()
        cursor.execute(query_str, args)
        try:
            if many:
                return cursor.fetchall()
            return cursor.fetchone()
        except Exception as error:
            return None


def import_data(label):
    file_path, _ = QFileDialog.getOpenFileName(
        None, "Выберите файл Excel", "", "Excel files (*xlsx)"
    )
    if file_path:
        wb = load_workbook(file_path)
        sheet = wb.active
        for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if i == 1:
                continue
            name = row[1]
            quantity = row[2]
            price = row[3]
            query(
                "INSERT INTO food (title, count, price) VALUES (?, ?, ?)",
                (name, quantity, price),
            )
        wb.close()
        label.setText("Все данные успешно импортированы")


def export_data(label):
    file_path, _ = QFileDialog.getSaveFileName(
        None, "Выберите файл Excel", "", "Excel files (*.xlsx)"
    )
    if file_path:
        wb = Workbook()
        sheet = wb.active
        sheet.append(["ID", "Title", "Count", "Price"])

        data = query("SELECT id, title, count, price FROM food ORDER BY id ASC")
        if data:
            for row in data:
                sheet.append(row)
            wb.save(file_path)
            wb.close()
            label.setText("Все данные успешно экспортированы")
    


def create_db():
    query(
        """
        CREATE TABLE IF NOT EXISTS food (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            count INTEGER NOT NULL,
            price REAL NOT NULL
        )
    """
    )


def drop_db():
    query("DROP TABLE IF EXISTS food")


def main():
    app = QApplication(sys.argv)  
    create_db()

    window = QMainWindow()
    window.setWindowTitle("DB to Excel")
    window.setGeometry(QRect(1280, 720, 300, 100))

    label = QLabel(window)
    label.setText("Для экспорта данных используйте пустой xlsx файл")

    import_button = QPushButton("Импортировать", window)
    import_button.clicked.connect(lambda: import_data(label))

    export_button = QPushButton("Экспортировать", window)
    export_button.clicked.connect(lambda: export_data(label))

    layout = QGridLayout()
    layout.addWidget(label, 0, 0, 1, 2)
    layout.addWidget(import_button, 1, 0)
    layout.addWidget(export_button, 1, 1)

    central_widget = QWidget()
    central_widget.setLayout(layout)
    window.setCentralWidget(central_widget)

    window.show()
    sys.exit(app.exec())
    # drop_db() # опасная штука за нее ручки могут оторваться
    # create_db()


if __name__ == "__main__":
    main()
